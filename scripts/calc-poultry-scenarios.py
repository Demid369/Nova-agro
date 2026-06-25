#!/usr/bin/env python3
"""Расчёт сценариев птицы: sensitivity, APK 100 млрд, экспорт, NPV."""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XLSX = Path(__file__).resolve().parents[1] / "docs/teo-poultry/_incoming/finmodel-pticekompleks-12-mlrd.xlsx"
OUT_JSON = Path(__file__).resolve().parents[1] / "docs/teo-poultry/appendix/calculated-scenarios.json"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

# --- Базовые объёмы (лист «Продукция») ---
MEAT_T = {
    "broiler": 17023.5,
    "turkey": 673.92,
    "duck": 252.0,
    "goose": 19.6992,
    "quail": 199.626,
}
EGGS_CHICKEN_MLN = 153.6
EGGS_QUAIL_MLN = 7.0
FERTILIZER_T = 43800
SIDE_MLN = 25.0

PRICES = {
    "broiler": 205,
    "turkey": 360,
    "duck": 320,
    "goose": 480,
    "quail": 450,
    "egg_chicken": 7.0,
    "egg_quail": 3.0,
    "fertilizer_kg": 12.0,
}

FEED_T = 66711.91
FEED_PRICE = 28000
FIXED_OPEX = 1475.0  # всё кроме корма (3342.93 - 1867.93)
CAPEX = 12000
LOADS = [0, 0.35, 0.65, 0.85, 1, 1, 1, 1, 1, 1]
RATE = 0.10


@dataclass
class Case:
    name: str
    revenue_mln: float
    opex_mln: float
    ebitda_mln: float
    margin_pct: float
    payback_yr: float
    npv_16y_mln: float
    irr_16y_pct: float
    note: str = ""


def revenue_mln(prices: dict | None = None) -> float:
    p = {**PRICES, **(prices or {})}
    total = SIDE_MLN + FERTILIZER_T * p["fertilizer_kg"] / 1000
    total += MEAT_T["broiler"] * p["broiler"] / 1000
    total += MEAT_T["turkey"] * p["turkey"] / 1000
    total += MEAT_T["duck"] * p["duck"] / 1000
    total += MEAT_T["goose"] * p["goose"] / 1000
    total += MEAT_T["quail"] * p["quail"] / 1000
    total += EGGS_CHICKEN_MLN * p["egg_chicken"]
    total += EGGS_QUAIL_MLN * p["egg_quail"]
    return round(total, 2)


def opex_mln(feed_price: float = FEED_PRICE, load: float = 1.0) -> float:
    feed = FEED_T * feed_price / 1e6 * load  # корм масштабируется с загрузкой
    # часть fixed opex тоже масштабируется при load < 1 (упрощённо 90% variable portion on feed+packaging)
    variable_extra = 280 + 190  # молодняк+упаковка грубо
    fixed_part = FIXED_OPEX - variable_extra
    return round(feed + fixed_part + variable_extra * load, 2)


def npv_irr(ebitda_full: float, capex: float = CAPEX, years_extra: int = 6):
    amort = capex / 10
    fcfs = []
    for i, load in enumerate(LOADS):
        ebitda = ebitda_full * load
        amort_y = amort * load
        profit = ebitda - amort_y
        tax = max(profit * 0.06, 0)
        fcf = profit - tax + amort_y
        if i == 0:
            fcf -= capex
        fcfs.append(fcf)
    # extend full load
    full_fcf = fcfs[-1]
    while len(fcfs) < 10 + years_extra:
        fcfs.append(full_fcf)
    npv = sum(f / (1 + RATE) ** i for i, f in enumerate(fcfs))
    # IRR bisection
    lo, hi = -0.3, 0.5
    for _ in range(80):
        mid = (lo + hi) / 2
        v = sum(f / (1 + mid) ** i for i, f in enumerate(fcfs))
        if v > 0:
            lo = mid
        else:
            hi = mid
    irr = (lo + hi) / 2 * 100
    return round(npv, 1), round(irr, 2)


def build_case(name: str, rev: float, opx: float, note: str = "") -> Case:
    eb = rev - opx
    npv, irr = npv_irr(eb)
    return Case(
        name=name,
        revenue_mln=rev,
        opex_mln=opx,
        ebitda_mln=round(eb, 2),
        margin_pct=round(eb / rev * 100, 1) if rev else 0,
        payback_yr=round(CAPEX / eb, 2) if eb > 0 else 999,
        npv_16y_mln=npv,
        irr_16y_pct=irr,
        note=note,
    )


def sensitivity_cases() -> list[Case]:
    base_rev = revenue_mln()
    base_opx = opex_mln()
    cases = [build_case("База (II кв. 2026)", base_rev, base_opx)]

    cases.append(
        build_case(
            "Яйцо −15%",
            revenue_mln({"egg_chicken": 5.95}),
            base_opx,
            "153,6 млн шт × 5,95 ₽",
        )
    )
    cases.append(
        build_case(
            "Яйцо −30%",
            revenue_mln({"egg_chicken": 4.9}),
            base_opx,
            "профицит рынка",
        )
    )
    cases.append(
        build_case(
            "Корм +15%",
            base_rev,
            opex_mln(32200),
            "28 000 → 32 200 ₽/т",
        )
    )
    cases.append(
        build_case(
            "Корм +20%",
            base_rev,
            opex_mln(33600),
            "стресс сырья",
        )
    )
    cases.append(
        build_case(
            "Бройлер −10%",
            revenue_mln({"broiler": 184.5}),
            base_opx,
            "205 → 184,5 ₽/кг",
        )
    )
    cases.append(
        build_case(
            "Ramp 85% (как 2029)",
            round(base_rev * 0.85, 2),
            round(base_opx * 0.88, 2),
            "выручка ×0,85; OPEX ×0,88 (часть постоянных)",
        )
    )
    cases.append(
        build_case(
            "Комбинированный стресс",
            revenue_mln({"egg_chicken": 5.6, "broiler": 195}),
            opex_mln(32200),
            "яйцо −20%, бройлер −5%, корм +15%",
        )
    )
    cases.append(
        build_case(
            "Оптимист (+5% цены мясо+яйцо)",
            revenue_mln({k: PRICES[k] * 1.05 for k in ("broiler", "turkey", "duck", "goose", "quail", "egg_chicken")}),
            base_opx,
        )
    )
    return cases


def export_cases() -> list[dict]:
    base = revenue_mln()
    meat_rev = (
        MEAT_T["broiler"] * PRICES["broiler"] / 1000
        + MEAT_T["turkey"] * PRICES["turkey"] / 1000
        + MEAT_T["duck"] * PRICES["duck"] / 1000
        + MEAT_T["goose"] * PRICES["goose"] / 1000
        + MEAT_T["quail"] * PRICES["quail"] / 1000
    )
    opx = opex_mln()
    export_opex_add = 11.0  # логистика, FTE, сертификация (export-plan-friendly-countries.md)
    rows = []
    scenarios = [
        ("Экспорт 0% (база finmodel)", 0, 0, 0, "весь сбыт РФ"),
        (
            "План 6% мяса, дружественные страны",
            round(meat_rev * 0.06, 1),
            0.02,
            export_opex_add,
            "235 млн ₽; KZ/UZ/BY/IR/KG/AE/AM/EG/AZ; год 4+",
        ),
        ("Экспорт 10% выручки (стресс-тест)", round(base * 0.10, 1), 0.03, 18, "агрессивный; не базовый план"),
        (
            "Экспорт 15% только мясо (стресс-тест)",
            round(meat_rev * 0.15, 1),
            0.03,
            22,
            "верхняя граница при перепроизводстве",
        ),
    ]
    for name, export_mln, premium_rate, opex_add, note in scenarios:
        premium = export_mln * premium_rate if export_mln else 0
        rev = base + premium
        eb = rev - opx - (opex_add if export_mln else 0)
        npv, irr = npv_irr(eb)
        rows.append(
            {
                "scenario": name,
                "export_mln_rub": export_mln,
                "revenue_mln": round(rev, 1),
                "ebitda_mln": round(eb, 1),
                "npv_16y_mln": npv,
                "irr_pct": irr,
                "note": note,
            }
        )
    return rows


def apk_100_scenarios() -> list[dict]:
    blocks = [
        ("Птица / кролики (блок 1)", 11041.5, 12000),
        ("Теплицы", 34500, 34500),
        ("Животноводство", 27458.5, 27458.5),
        ("Рыба", 9000, 9000),
        ("Масложир", 18000, 18000),
    ]
    scenarios = []

    # A: replace as-is
    total_a = sum(b[2] for b in blocks)
    scenarios.append({"name": "A: замена 1:1 (+958,5)", "total_mln": total_a, "delta": total_a - 100000})

    # B: poultry trimmed to rabbit capex
    total_b = 11041.5 + sum(b[2] for b in blocks[1:])
    scenarios.append({"name": "B: птица = 11 041,5 (как кролики)", "total_mln": total_b, "delta": 0, "poultry_capex": 11041.5})

    # C: proportional trim 958.5 from blocks 2-5
    trim = 958.5
    others = sum(b[1] for b in blocks[1:])
    trimmed = [12000]
    for _, base, _ in blocks[1:]:
        cut = trim * (base / others)
        trimmed.append(round(base - cut, 1))
    total_c = sum(trimmed)
    scenarios.append(
        {
            "name": "C: птица 12 000, остальные −958,5 пропорц.",
            "total_mln": round(total_c, 1),
            "delta": round(total_c - 100000, 1),
            "blocks_mln": dict(zip([b[0] for b in blocks], trimmed)),
        }
    )
    return scenarios


def write_xlsx_sheets(cases: list[Case], export_rows: list[dict], apk_rows: list[dict]) -> None:
    wb = openpyxl.load_workbook(XLSX)

    if "Чувствительность" in wb.sheetnames:
        del wb["Чувствительность"]
    ws = wb.create_sheet("Чувствительность")
    ws["A1"] = "АНАЛИЗ ЧУВСТВИТЕЛЬНОСТИ — полная мощность"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "NPV/IRR: DCF 16 лет @10%, как лист NPV-IRR · расчёт scripts/calc-poultry-scenarios.py"
    headers = ["Сценарий", "Выручка", "OPEX", "EBITDA", "Маржа %", "PB лет", "NPV @10%", "IRR %", "Комментарий"]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i, h)
        ws.cell(4, i).font = HEADER_FONT
        ws.cell(4, i).fill = HEADER_FILL
    for i, c in enumerate(cases):
        r = 5 + i
        ws.cell(r, 1, c.name)
        ws.cell(r, 2, c.revenue_mln)
        ws.cell(r, 3, c.opex_mln)
        ws.cell(r, 4, c.ebitda_mln)
        ws.cell(r, 5, c.margin_pct / 100)
        ws.cell(r, 5).number_format = "0.0%"
        ws.cell(r, 6, c.payback_yr)
        ws.cell(r, 7, c.npv_16y_mln)
        ws.cell(r, 8, c.irr_16y_pct / 100)
        ws.cell(r, 8).number_format = "0.0%"
        ws.cell(r, 9, c.note)

    if "Экспорт" in wb.sheetnames:
        del wb["Экспорт"]
    ws2 = wb.create_sheet("Экспорт")
    ws2["A1"] = "СЦЕНАРИИ ЭКСПОРТА (гипотезы)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A2"] = "Премия к выручке упрощённо: +3% на экспортный объём (логистика/FX не моделируются)"
    h2 = ["Сценарий", "Экспорт, млн ₽", "Выручка итого", "EBITDA", "NPV @10%", "IRR %", "Примечание"]
    for i, h in enumerate(h2, 1):
        ws2.cell(4, i, h)
        ws2.cell(4, i).font = HEADER_FONT
    for i, row in enumerate(export_rows):
        r = 5 + i
        ws2.cell(r, 1, row["scenario"])
        ws2.cell(r, 2, row["export_mln_rub"])
        ws2.cell(r, 3, row["revenue_mln"])
        ws2.cell(r, 4, row["ebitda_mln"])
        ws2.cell(r, 5, row["npv_16y_mln"])
        ws2.cell(r, 6, row["irr_pct"] / 100)
        ws2.cell(r, 6).number_format = "0.0%"
        ws2.cell(r, 7, row["note"])

    # Extend APK-100 with scenario C block if sheet exists
    if "APK-100" in wb.sheetnames:
        ws3 = wb["APK-100"]
        ws3["A26"] = "Сценарий C: итог ровно 100 000"
        ws3["A26"].font = HEADER_FONT
        if apk_rows and "blocks_mln" in apk_rows[2]:
            for i, (name, val) in enumerate(apk_rows[2]["blocks_mln"].items()):
                ws3.cell(27 + i, 1, name)
                ws3.cell(27 + i, 2, val)
            ws3.cell(32, 1, "ИТОГО")
            ws3.cell(32, 2, apk_rows[2]["total_mln"])

    wb.save(XLSX)


def main() -> None:
    cases = sensitivity_cases()
    export_rows = export_cases()
    apk_rows = apk_100_scenarios()

    payload = {
        "base": cases[0].__dict__,
        "sensitivity": [c.__dict__ for c in cases],
        "export": export_rows,
        "apk_100": apk_rows,
    }
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx_sheets(cases, export_rows, apk_rows)

    print("=== БАЗА ===")
    b = cases[0]
    print(f"Rev {b.revenue_mln} OPEX {b.opex_mln} EBITDA {b.ebitda_mln} NPV {b.npv_16y_mln} IRR {b.irr_16y_pct}%")

    print("\n=== ЧУВСТВИТЕЛЬНОСТЬ ===")
    for c in cases[1:]:
        print(f"{c.name}: EBITDA {c.ebitda_mln} NPV {c.npv_16y_mln} IRR {c.irr_16y_pct}%")

    print("\n=== ЭКСПОРТ ===")
    for r in export_rows:
        print(r)

    print("\n=== APK 100 млрд ===")
    for r in apk_rows:
        print(r)

    print(f"\nSaved: {XLSX}, {OUT_JSON}")


if __name__ == "__main__":
    main()
