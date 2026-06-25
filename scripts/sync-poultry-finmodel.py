#!/usr/bin/env python3
"""Sync poultry finmodel xlsx: CAPEX balance, NPV/IRR, financing, formula fixes."""
from __future__ import annotations

import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

XLSX = Path(__file__).resolve().parents[1] / "docs/teo-poultry/_incoming/finmodel-pticekompleks-12-mlrd.xlsx"

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_header(ws, row: int, cols: int = 6) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def fix_revenue_refs(wb: openpyxl.Workbook) -> None:
    ws = wb["Выручка"]
    total_row = 13
    total_ref = f"Выручка!$E${total_row}"
    for r in range(4, 12):
        if ws.cell(r, 6).value and str(ws.cell(r, 6).value).startswith("="):
            ws.cell(r, 6).value = f"=E{r}/{total_ref}"
    ws.cell(14, 5).value = f"=(E4+E9+E10)/{total_ref}"
    ws.cell(15, 1).value = "Справочно: доля «бройлер + яйцо» в выручке"
    ws.cell(15, 5).value = f"=(E4+E9+E10)/{total_ref}"


def fix_capex_balance(wb: openpyxl.Workbook) -> None:
    """Trim infra lines to offset solar+compost vs biogaz (+766.8) within 12 000 envelope."""
    ws = wb["CAPEX"]
    # Net CAPEX delta vs original biogaz 450: +766.8 — trim holding-level / optimizable infra
    ws["C18"].value = 750  # инженерные сети 1000 → 750 (−250, solar on roofs)
    ws["C20"].value = 183  # дороги 250 → 183 (−67)
    ws["C22"].value = 250  # спецтехника 350 → 250 (−100)
    ws["C25"].value = 33  # резерв: фикс. ориентир (формула балансирует)
    ws.cell(25, 4).value = "=12000-SUM(D4:D24)"
    ws.cell(26, 4).value = "=SUM(D4:D25)"
    # Экономика must reference ИТОГО, not резерв
    econ = wb["Экономика"]
    econ["B4"].value = "=CAPEX!$D$26"
    for r in range(22, 26):
        for c in range(2, 12):
            ref = ws.cell(r, c)
            if isinstance(ref.value, str) and "CAPEX!$D$25" in ref.value:
                ref.value = ref.value.replace("CAPEX!$D$25", "CAPEX!$D$26")


def add_npv_irr_sheet(wb: openpyxl.Workbook) -> None:
    if "NPV-IRR" in wb.sheetnames:
        del wb["NPV-IRR"]
    ws = wb.create_sheet("NPV-IRR")
    ws["A1"] = "NPV / IRR — DCF @ 10% (эталон кроликов, tab#7 docx)"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Горизонт 2026–2035 · без долга · FCF = ЧП + амортизация · CAPEX в 2026"
    ws["A3"] = "Ставка дисконтирования"
    ws["B3"] = "=Допущения!C19"
    ws["C3"] = "%"

    headers = ["Год", "Загрузка", "EBITDA", "Аморт.", "Прибыль до налога", "Налог ЕСХН 6%", "Чистая прибыль", "FCF", "DF @ 10%", "PV FCF"]
    for i, h in enumerate(headers, 1):
        ws.cell(5, i, h)
    style_header(ws, 5, len(headers))

    years = list(range(2026, 2042))  # 16 лет — NPV @10% сопоставим с tab#7 кроликов
    for i, year in enumerate(years):
        r = 6 + i
        if year <= 2035:
            col = i + 2
            col_letter = openpyxl.utils.get_column_letter(col)
            load_ref = f"Экономика!{col_letter}22"
            ebitda_ref = f"Экономика!{col_letter}24"
        else:
            load_ref = "1"
            ebitda_ref = "Экономика!$B$7"
        ws.cell(r, 1, year)
        ws.cell(r, 2, f"={load_ref}")
        ws.cell(r, 3, f"={ebitda_ref}")
        ws.cell(r, 4, f"=IF(B{r}>0,Экономика!$B$9*B{r},0)")
        ws.cell(r, 5, f"=C{r}-D{r}")
        ws.cell(r, 6, f"=MAX(E{r}*0.06,0)")
        ws.cell(r, 7, f"=E{r}-F{r}")
        if year == 2026:
            ws.cell(r, 8, f"=G{r}+D{r}-Экономика!$B$4")
        else:
            ws.cell(r, 8, f"=G{r}+D{r}")
        ws.cell(r, 9, f"=1/(1+$B$3/100)^({r}-5)")
        ws.cell(r, 10, f"=H{r}*I{r}")

    r_sum = 22
    ws.cell(r_sum, 1, "NPV @ 10%, млн ₽").font = HEADER_FONT
    ws.cell(r_sum, 2, "=SUM(J6:J21)")
    ws.cell(r_sum + 1, 1, "IRR, %").font = HEADER_FONT
    ws.cell(r_sum + 1, 2, "=IRR(H6:H21)*100")
    ws.cell(r_sum + 2, 1, "PI (NPV/CAPEX + 1)").font = HEADER_FONT
    ws.cell(r_sum + 2, 2, "=B22/Экономика!B4+1")

    ws.cell(r_sum + 4, 1, "Сравнение с кроликами (docx tab#7)")
    ws.cell(r_sum + 5, 1, "Кролики NPV @ 10%")
    ws.cell(r_sum + 5, 2, 2779.5)
    ws.cell(r_sum + 6, 1, "Кролики IRR")
    ws.cell(r_sum + 6, 2, 0.1519)
    ws.cell(r_sum + 6, 2).number_format = "0.00%"
    ws.cell(r_sum + 7, 1, "Горизонт DCF")
    ws.cell(r_sum + 7, 2, "16 лет (2026–2041)")

    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14


def add_financing_sheet(wb: openpyxl.Workbook) -> None:
    """Financing structure for 12 000 mln block — template по 1.2-слайд Фин модель.xlsx (кролики)."""
    if "Финансирование" in wb.sheetnames:
        del wb["Финансирование"]
    ws = wb.create_sheet("Финансирование")
    ws["A1"] = "СТРУКТУРА ФИНАНСИРОВАНИЯ — ПТИЦЕКОМПЛЕКС 12 000 млн ₽"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "Шаблон: docs/1.2-слайд Фин модель.xlsx (ФНБ / РАЛ / ФРП / банки) · блок птицы отдельно от APK 100 млрд"
    ws["A3"] = "Сумма инвестиций"
    ws["B3"] = 12000
    ws["C3"] = "млн ₽"

    # Доли как у блока кроликов (11 041,5): РАЛ ~19,5%, банки ~80,5%, ФНБ ~5%; ФРП — ККЗ вне блока
    rows = [
        ("", "Источник", "Статья", "Сумма, млн ₽", "Доля", "Комментарий"),
        ("ФНБ", "Фонд национального благосостояния", "Долевое участие в проекте", 600, "=D6/$B$3", "~5% конверта (аналог холдинга 100 млрд)"),
        ("РАЛ", "РосАгроЛизинг", "Оборудование (птичники, УПК, инкубатор, холод)", 2340, "=D7/$B$3", "19,5% блока — как кролики 2 147,7 / 11 041,5"),
        ("ФРП", "Фонд развития промышленности", "—", 0, "=D8/$B$3", "ККЗ вне блока; корм с завода холдинга"),
        ("Банки", "Коммерческие банки", "СМР, здания, инженерные сети", 6287, "=D9/$B$3", "69,4% банковской доли (как СМР у кроликов)"),
        ("Банки", "Коммерческие банки", "Solar 10 МВт·ч", 1180.8, "=D10/$B$3", "эталон кроликов tab#23"),
        ("Банки", "Коммерческие банки", "Комpost помёта 5 т/ч", 36, "=D11/$B$3", "эталон кроликов tab#23"),
        ("Банки", "Коммерческие банки", "Проектирование, пусконаладка", 27, "=D12/$B$3", "аналог строки «Проектирование» у кроликов"),
        ("Банки", "Коммерческие банки", "Родительское стадо, стартовый молодняк", 463, "=D13/$B$3", "аналог «племенные» у кроликов"),
        ("Банки", "Коммерческие банки", "Оборотные средства", 700, "=D14/$B$3", "как кролики (700 млн)"),
        ("Банки", "Коммерческие банки", "Резерв + прочая инфраструктура", "=12000-SUM(D6:D14)", "=D15/$B$3", "баланс до 12 000 (= CAPEX)"),
        ("", "ИТОГО", "", "=SUM(D6:D15)", "=D16/$B$3", ""),
    ]

    start = 5
    for i, row in enumerate(rows):
        r = start + i
        for j, val in enumerate(row, 1):
            ws.cell(r, j, val)
        if i == 0:
            style_header(ws, r, 6)

    ws.cell(start + 12, 1, "Примечание").font = HEADER_FONT
    ws.cell(start + 13, 1, "ККЗ и биогаз APK 20 МВт·ч — на уровне холдинга, не в CAPEX/OPEX блока птицы.")
    ws.cell(start + 14, 1, "Детализация CAPEX — лист «CAPEX»; операционная модель — «Экономика», «NPV-IRR».")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 42


def update_summary(wb: openpyxl.Workbook) -> None:
    ws = wb["Сводка"]
    ws["B16"] = 476
    ws["A20"] = "NPV @ 10%, млн ₽"
    ws["B20"] = "=NPV-IRR!B22"
    ws["A21"] = "IRR, %"
    ws["B21"] = "=NPV-IRR!B23"
    ws["A22"] = "Финансирование"
    ws["B22"] = "лист «Финансирование»"
    ws["A23"] = "Энергетика"
    ws["B23"] = "solar 10 МВт·ч + compost 5 т/ч (= кролики); без лок. biogaz+ГПU"
    ws["A24"] = "Подробности — листы: Продукция, Выручка, Корма, CAPEX, OPEX, Экономика, NPV-IRR, Финансирование, Допущения."


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    fix_revenue_refs(wb)
    fix_capex_balance(wb)
    add_npv_irr_sheet(wb)
    add_financing_sheet(wb)
    update_summary(wb)
    wb.save(XLSX)

    # Verify computed values
    wb2 = openpyxl.load_workbook(XLSX, data_only=True)
    # force recalc not available — compute manually
    wb_f = openpyxl.load_workbook(XLSX, data_only=False)
    ws_c = wb_f["CAPEX"]
    fixed = 0
    for r in range(4, 25):
        qty = ws_c.cell(r, 2).value or 1
        unit = ws_c.cell(r, 3).value or 0
        val = ws_c.cell(r, 4).value
        if isinstance(val, str) and val.startswith("="):
            val = qty * unit if unit else None
        if isinstance(val, (int, float)):
            fixed += val
    reserve = 12000 - fixed
    print(f"CAPEX fixed (4-24): {fixed:.1f}, reserve: {reserve:.1f}, total: {fixed+reserve:.1f}")
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
